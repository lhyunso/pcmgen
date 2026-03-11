"""
Excel Export for PCM Channel Table - IRIG 106 compliant output.
Generates professional telemetry channel allocation spreadsheets.
3 sheets: Summary, Parameters, Channel Table
"""
import math
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter


def hex_to_rgb(hex_color: str) -> str:
    """Convert #RRGGBB to RRGGBB for openpyxl."""
    return hex_color.lstrip("#")


def is_dark_bg(hex_color: str) -> bool:
    """Determine if text should be white on this background."""
    h = hex_color.lstrip("#")
    if len(h) != 6:
        return False
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return (r * 299 + g * 587 + b * 114) / 1000 < 160


# ── Styles ──
HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(name="맑은 고딕", size=9, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="맑은 고딕", size=14, bold=True, color="1E3A5F")
SUBTITLE_FONT = Font(name="맑은 고딕", size=10, bold=True, color="4A5568")
SECTION_FONT = Font(name="맑은 고딕", size=11, bold=True, color="1A6FB5")
NORMAL_FONT = Font(name="맑은 고딕", size=9)
BOLD_FONT = Font(name="맑은 고딕", size=9, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ROW_LABEL_FILL = PatternFill("solid", fgColor="E2E8F0")
RSVD_FILL = PatternFill("solid", fgColor="F5F5F5")


def export_to_excel(config: dict, result: dict, filepath: str) -> str:
    """Export PCM channel table to Excel workbook."""
    wb = Workbook()

    _write_summary_sheet(wb, config, result)
    _write_param_sheet(wb, config, result)
    _write_channel_table_sheet(wb, config, result)

    wb.save(filepath)
    return filepath


def _styled_cell(ws, row, col, value, font=NORMAL_FONT, fill=None,
                 alignment=LEFT, border=THIN_BORDER):
    """Write a styled cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = alignment
    cell.border = border
    return cell


def _write_summary_sheet(wb: Workbook, config: dict, result: dict):
    """Sheet 1: Configuration Summary."""
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "1E3A5F"

    stats = result["stats"]
    row = 1

    # Title
    ws.merge_cells("A1:D1")
    _styled_cell(ws, 1, 1, "PCM Channel Configuration Summary",
                 font=TITLE_FONT, alignment=Alignment(horizontal="center", vertical="center"))
    row = 2
    ws.merge_cells("A2:D2")
    _styled_cell(ws, 2, 1, "IRIG 106 Chapter 4 Compliant",
                 font=SUBTITLE_FONT, alignment=Alignment(horizontal="center"))
    row = 3
    ws.merge_cells("A3:D3")
    _styled_cell(ws, 3, 1, "\u00a9 2025 단암시스템즈 (Danam Systems)",
                 font=Font(name="맑은 고딕", size=9, color="718096"),
                 alignment=Alignment(horizontal="center"))
    row = 4

    # Frame Parameters
    _styled_cell(ws, row, 1, "Frame Parameters", font=SECTION_FONT)
    row += 1

    frame_data = [
        ("PCM Class", stats["pcm_class"]),
        ("Bit Rate", f"{stats['bitrate']:,} bps ({stats['bitrate_mbps']} Mbps)"),
        ("Bit Representation", stats["bit_repr"]),
        ("Words / Minor Frame", f"{stats['words_per_minor']:,}"),
        ("Bits / Word", str(stats["bits_per_word"])),
        ("Minor Frames (Z)", str(stats["minor_count"])),
        ("Major Frame Rate", f"{stats['frame_hz']} Hz"),
        ("Minor Frame Rate", f"{stats.get('minor_frame_rate', stats['frame_hz'])} Hz"),
        ("SYNC Pattern", stats["sync_pattern"]),
        ("SYNC Bits", f"{stats['sync_bits']} bit"),
        ("Overhead / Minor", f"{stats['overhead_per_minor']} words (SYNC + SFID)"),
    ]

    for label, value in frame_data:
        _styled_cell(ws, row, 1, label, font=BOLD_FONT)
        _styled_cell(ws, row, 2, value)
        row += 1

    row += 1

    # Channel Usage
    _styled_cell(ws, row, 1, "Channel Usage", font=SECTION_FONT)
    row += 1

    usage_data = [
        ("Total Available", f"{stats['total_available']:,} words"),
        ("Sensor", f"{stats['sensor_ch']:,} words"),
        ("Digital", f"{stats['digital_ch']:,} words"),
        ("BIT", f"{stats['bit_ch']:,} words"),
        ("Total Used", f"{stats['total_used']:,} words"),
        ("Reserved (Spare)", f"{stats['total_reserved']:,} words"),
        ("Usage", f"{stats['usage_pct']}%"),
        ("Max Supercommutation", f"{stats['max_supercom']}x"),
    ]

    if stats.get("subcom_positions", 0) > 0:
        usage_data.append(("Subcom Positions", f"{stats['subcom_positions']} words ({stats.get('subcom_params', 0)} params)"))

    for label, value in usage_data:
        _styled_cell(ws, row, 1, label, font=BOLD_FONT)
        _styled_cell(ws, row, 2, value)
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 45


def _write_param_sheet(wb: Workbook, config: dict, result: dict):
    """Sheet 2: Parameter List with Comm and Words/Minor."""
    ws = wb.create_sheet("Parameters")
    ws.sheet_properties.tabColor = "2D7D46"

    headers = ["#", "Category", "Type", "Name", "Hz", "EA/Bits",
               "Comm", "Words/Minor", "Note"]
    widths = [5, 10, 14, 24, 8, 10, 14, 14, 30]

    for ci, h in enumerate(headers, 1):
        _styled_cell(ws, 1, ci, h, font=HEADER_FONT, fill=HEADER_FILL,
                     alignment=CENTER)

    param_list = result.get("param_list", [])

    # If param_list is empty, build from config
    if not param_list:
        param_list = _build_param_list_from_config(config, result.get("stats", {}))

    for ri, p in enumerate(param_list, 2):
        vals = [
            p.get("idx", ri - 1),
            p.get("category", ""),
            p.get("type", ""),
            p.get("name", ""),
            p.get("hz", ""),
            p.get("ea", ""),
            p.get("comm", "1x"),
            p.get("words_per_minor", ""),
            p.get("note", ""),
        ]

        # Color the row based on sensor color
        color = p.get("color", "")
        row_fill = None
        row_font = NORMAL_FONT
        if color:
            rgb = hex_to_rgb(color)
            if len(rgb) == 6:
                # Use a lighter version for row background
                row_fill = PatternFill("solid", fgColor=rgb)
                if is_dark_bg(color):
                    row_font = Font(name="맑은 고딕", size=9, color="FFFFFF")

        for ci, v in enumerate(vals, 1):
            cell = _styled_cell(ws, ri, ci, v,
                                font=row_font if row_fill else NORMAL_FONT,
                                fill=row_fill,
                                alignment=CENTER if ci <= 8 else LEFT)

    # Totals row
    total_row = len(param_list) + 2
    _styled_cell(ws, total_row, 1, "TOTAL", font=BOLD_FONT,
                 fill=PatternFill("solid", fgColor="E2E8F0"), alignment=CENTER)
    total_wpm = sum(p.get("words_per_minor", 0) for p in param_list
                    if isinstance(p.get("words_per_minor"), (int, float)))
    _styled_cell(ws, total_row, 8, total_wpm, font=BOLD_FONT,
                 fill=PatternFill("solid", fgColor="E2E8F0"), alignment=CENTER)
    for ci in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=ci).border = THIN_BORDER

    # Column widths
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(param_list) + 1}"
    ws.freeze_panes = "A2"


def _build_param_list_from_config(config: dict, stats: dict):
    """Fallback: build param_list from config when JS didn't provide it."""
    param_list = []
    mfr = config.get("frame_hz", 50) * config.get("minor_count", 1)
    bits_per_word = config.get("bits_per_word", 16)
    word_bytes = bits_per_word / 8
    idx = 1

    for s in config.get("sensors", []):
        hz = s.get("hz", 50)
        ch = s.get("ch", 1)
        sc = hz / mfr if mfr > 0 else 1
        is_sub = sc < 1
        sc_r = 1 if is_sub else max(1, round(sc))
        sub_r = round(1 / sc) if is_sub else 0
        comm = f"Sub 1/{sub_r}" if is_sub else f"Super {sc_r}x" if sc_r > 1 else "1x"
        wpm = ch if is_sub else ch * sc_r
        param_list.append({
            "idx": idx, "category": "Sensor", "type": s.get("type", ""),
            "name": s.get("name", ""), "hz": hz, "ea": ch,
            "comm": comm, "words_per_minor": wpm, "note": s.get("note", ""),
        })
        idx += 1

    for d in config.get("digital", []):
        hz = d.get("hz", 50)
        bits = d.get("bits", 16)
        ch = math.ceil(bits / bits_per_word)
        sc = hz / mfr if mfr > 0 else 1
        is_sub = sc < 1
        sc_r = 1 if is_sub else max(1, round(sc))
        sub_r = round(1 / sc) if is_sub else 0
        comm = f"Sub 1/{sub_r}" if is_sub else f"Super {sc_r}x" if sc_r > 1 else "1x"
        wpm = ch if is_sub else ch * sc_r
        param_list.append({
            "idx": idx, "category": "Digital", "type": d.get("type", ""),
            "name": d.get("name", ""), "hz": hz, "ea": f"{bits}bits",
            "comm": comm, "words_per_minor": wpm, "note": d.get("note", ""),
        })
        idx += 1

    for b in config.get("bit_data", []):
        byt = b.get("bytes", 0)
        ch = math.ceil(byt / word_bytes) if word_bytes > 0 else 0
        param_list.append({
            "idx": idx, "category": "BIT", "type": "VL",
            "name": f"VL#{b.get('vl', idx)}", "hz": "-",
            "ea": f"{byt}bytes", "comm": "Sub",
            "words_per_minor": ch, "note": b.get("note", ""),
        })
        idx += 1

    return param_list


def _write_channel_table_sheet(wb: Workbook, config: dict, result: dict):
    """Sheet 3: Full Channel Allocation Table (W columns x Z rows)."""
    ws = wb.create_sheet("Channel Table")
    ws.sheet_properties.tabColor = "B45309"

    frames = result.get("frames", [])
    if not frames:
        ws["A1"] = "No channel table generated"
        return

    W = config.get("words_per_minor", len(frames[0].get("cells", [])))
    Z = len(frames)

    # Header row: Word#
    _styled_cell(ws, 1, 1, "SFID \\ Word#", font=HEADER_FONT, fill=HEADER_FILL,
                 alignment=CENTER)

    for ci in range(W):
        cell = ws.cell(row=1, column=ci + 2, value=ci + 1)
        cell.font = Font(name="맑은 고딕", size=7, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Data rows
    for ri, frame_data in enumerate(frames, 2):
        # Row label
        _styled_cell(ws, ri, 1, frame_data.get("label", f"MF {ri - 1}"),
                     font=BOLD_FONT, fill=ROW_LABEL_FILL, alignment=CENTER)

        cells = frame_data.get("cells", [])
        for ci in range(min(W, len(cells))):
            c = cells[ci]
            cell = ws.cell(row=ri, column=ci + 2)

            name = c.get("name", "")
            is_rsvd = name == "RSVD" or name == ""

            if is_rsvd:
                cell.value = ""
                cell.fill = RSVD_FILL
            else:
                cell.value = name
                color = c.get("color", "#F5F5F5")
                rgb = hex_to_rgb(color)
                if len(rgb) == 6:
                    cell.fill = PatternFill("solid", fgColor=rgb)
                    if is_dark_bg(color):
                        cell.font = Font(name="맑은 고딕", size=6, color="FFFFFF")
                    else:
                        cell.font = Font(name="맑은 고딕", size=6, color="222222")

                # Subcom indicator
                if c.get("subcom"):
                    cell.font = Font(name="맑은 고딕", size=6, bold=True,
                                     color="FFFFFF" if is_dark_bg(color) else "E65100")

            cell.alignment = CENTER
            cell.border = THIN_BORDER

    # Column widths
    ws.column_dimensions["A"].width = 16
    for ci in range(W):
        col_letter = get_column_letter(ci + 2)
        ws.column_dimensions[col_letter].width = 4

    # Freeze panes
    ws.freeze_panes = "B2"

    # Legend from result data
    legend_items = result.get("legend", [])
    if not legend_items:
        # Fallback hardcoded legend
        legend_items = [
            {"label": "SYNC", "color": "#FF4444"},
            {"label": "SFID", "color": "#CC0000"},
            {"label": "RSVD", "color": "#F5F5F5"},
        ]

    legend_row = Z + 3
    _styled_cell(ws, legend_row, 1, "Legend:", font=BOLD_FONT)
    for ci, item in enumerate(legend_items):
        label = item.get("label", "")
        color = item.get("color", "#F5F5F5")
        rgb = hex_to_rgb(color)
        cell = ws.cell(row=legend_row + 1, column=ci + 1, value=label)
        if len(rgb) == 6:
            cell.fill = PatternFill("solid", fgColor=rgb)
        if is_dark_bg(color):
            cell.font = Font(name="맑은 고딕", size=8, color="FFFFFF")
        else:
            cell.font = Font(name="맑은 고딕", size=8, color="222222")
        cell.alignment = CENTER
        cell.border = THIN_BORDER
